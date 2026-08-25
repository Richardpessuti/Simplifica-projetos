"""Lê e grava o histórico de preços em uma planilha .xlsx (pra abrir no Excel)
e em um .json (pra alimentar a página de preços do site)."""
import datetime
import json
from pathlib import Path

import openpyxl

COLUNAS = ["data", "produto", "site", "nome_encontrado", "preco", "url", "erro"]
ABA = "Histórico"


def _abrir_ou_criar(caminho: Path):
    if caminho.exists():
        return openpyxl.load_workbook(caminho)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = ABA
    ws.append(COLUNAS)
    return wb


def salvar_resultados(caminho: Path, resultados):
    caminho.parent.mkdir(parents=True, exist_ok=True)
    wb = _abrir_ou_criar(caminho)
    ws = wb[ABA]
    for r in resultados:
        ws.append([r.get(c, "") for c in COLUNAS])
    wb.save(caminho)


def salvar_json(caminho: Path, resultados):
    """Acrescenta os resultados de hoje em data/historico.json (lista simples de
    linhas), usado pela página precos.html — mais fácil de ler via fetch()
    no navegador do que abrir o .xlsx."""
    caminho.parent.mkdir(parents=True, exist_ok=True)
    linhas = []
    if caminho.exists():
        linhas = json.loads(caminho.read_text(encoding="utf-8")).get("linhas", [])
    for r in resultados:
        linhas.append({c: r.get(c) for c in COLUNAS})
    dados = {
        "atualizado_em": datetime.datetime.now(datetime.timezone.utc).isoformat(),
        "linhas": linhas,
    }
    caminho.write_text(json.dumps(dados, ensure_ascii=False, indent=2), encoding="utf-8")


def preco_anterior(caminho: Path, produto, site):
    """Último preço registrado pra esse produto+site antes de hoje (None se não houver)."""
    if not caminho.exists():
        return None
    wb = openpyxl.load_workbook(caminho)
    ws = wb[ABA]
    ultimo = None
    for linha in ws.iter_rows(min_row=2, values_only=True):
        _data, produto_linha, site_linha, _nome, preco, *_resto = linha
        if produto_linha == produto and site_linha == site and preco not in (None, ""):
            ultimo = preco
    return ultimo
